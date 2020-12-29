from django.contrib.auth.views import LoginView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.conf import settings
from django.core.files.base import ContentFile
from django.http import HttpResponseRedirect
from django.shortcuts import redirect
from django.db import transaction
from django.utils import timezone
from django.urls import reverse
from django.views.generic import ListView, DetailView, CreateView, UpdateView
from .models import Staff, Customer, GeneratedData, MatchedData, VisuallyMatchedData, ImportData, BillingData, Product, Agent, Place
from .forms import CustomerSelectAndFileUpLoadMultiFrom, VisuallyMatchedDataCreateForm, ImportDataCreateForm, \
    UploadFileSelectForm,  BillingDataFrom, BiilingFileUploadFrom, BillingDataFromSet
from datetime import datetime
import openpyxl,  os, csv, codecs, chardet, urllib.parse, re, io, math, pprint


UPLOAD_NOT_COMPLETED = 0
CSV_OUTPUT_COMPLETED = 1
VISUALLY_CONFIRMED = 2
IMPORT_DATA_OUTPUT_COMPLETED = 3


class LoginFormView(SuccessMessageMixin, LoginView):
    template_name = 'registration/login.html'

    def get_success_url(self):
        return reverse('top')

    def get_success_message(self, cleaned_data):
        return f'{self.request.user}でログインしました'

    def get_context_data(self, **kwargs):
        context = super().get_context_data()
        context['page'] = 'LOGIN'
        return context


from django.contrib.auth import logout as auth_logout

def logout(request):
    auth_logout(request)
    return redirect(reverse('login'))

class HistoryListView(LoginRequiredMixin, ListView):
    template_name = 'app/top.html'

    def get_queryset(self):
        return GeneratedData.objects.order_by('-status')

    def get_context_data(self, **kwargs):
        customer = Customer.objects.all()
        staff = Staff.objects.all()
        matched = MatchedData.objects.filter(created_date__lte=timezone.now()).order_by('-created_date')
        visually_matched = VisuallyMatchedData.objects.filter(created_date__lte=timezone.now()).order_by('-created_date')
        import_data = ImportData.objects.filter(created_date__lte=timezone.now()).order_by('-created_date')
        billing = BillingData.objects.all()

        context = super().get_context_data(**kwargs)
        context['customer'] = customer
        context['staff'] = staff
        context['matched'] = matched
        context['billing'] = billing
        context['visually_matched'] = visually_matched
        context['import_data'] = import_data
        context['page'] = 'TOP'

        return context


class CustomerSelectAndFileUpLoadView(LoginRequiredMixin, CreateView):
    form_class = CustomerSelectAndFileUpLoadMultiFrom
    template_name = 'app/file_upload.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status'] = UPLOAD_NOT_COMPLETED

        return context

    @transaction.atomic
    def form_valid(self, form):
        generated_data = form['generated_data'].save(commit=False)
        staff = Staff.objects.get(staff=self.request.user)
        generated_data.staff = staff
        generated_data.status = UPLOAD_NOT_COMPLETED
        generated_data.save()

        matched_data = form['matched_data'].save(commit=False)
        matched_data.generated = generated_data
        staff = Staff.objects.get(staff=self.request.user)
        matched_data.staff = staff
        matched_data.generated.status = CSV_OUTPUT_COMPLETED
        matched_data.save()

        brycen_file_path = urllib.parse.unquote(matched_data.brycen_file.path)
        billing_file_path = urllib.parse.unquote(matched_data.billing_file.path)
        print('brycen_file_path:{}'.format(brycen_file_path))
        print('billing_file_path:{}'.format(billing_file_path))

        # ファイル命名
        now = datetime.now()
        file_name = 'TXJ_付け合わせ済_' + now.strftime('%Y年%m月%d日%H時%M分%S秒') + '_作成分)' + '.csv'

        output_data = create_csv(brycen_file_path, billing_file_path)

        # ファイルsave
        matched_data.matched_data_file.save(file_name, ContentFile(output_data))

        return super().form_valid(form)

    def get_success_url(self):
        matched = self.object['matched_data']
        return reverse('detail_and_create', kwargs={'pk': matched.id})


class CustomerSelectAndBrycenFileUpLoadView(CreateView):
    form_class = CustomerSelectAndFileUpLoadMultiFrom
    template_name = 'app/customer_brycen_file_select.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['status'] = 'CREATE_START'

        return context

    def form_valid(self, form):
        generated_data = form['generated_data'].save(commit=False)
        staff = Staff.objects.get(staff=self.request.user)
        generated_data.staff = staff
        generated_data.status = UPLOAD_NOT_COMPLETED
        generated_data.save()

        matched_data = form['matched_data'].save(commit=False)
        matched_data.generated = generated_data
        staff = Staff.objects.get(staff=self.request.user)
        matched_data.staff = staff
        matched_data.generated.status = UPLOAD_NOT_COMPLETED
        matched_data.save()

        return super().form_valid(form)

    def get_success_url(self):
        matched = self.object['matched_data']
        return reverse('select_billing_file_or_form', kwargs={'pk': matched.id})


class SelectFileOrBillingDataFormView(CreateView):
    form_class = BiilingFileUploadFrom
    template_name = 'app/select_billing_file_or_form.html'

    def get_context_data(self, **kwargs):
        matched_data_pk = self.kwargs['pk']
        matched = MatchedData.objects.get(pk=matched_data_pk)
        billing = BillingData.objects.filter(matched_id=matched_data_pk)
        context = super().get_context_data(**kwargs)
        context['matched'] = matched
        context['billing'] = billing
        context['status'] = UPLOAD_NOT_COMPLETED

        return context

    def form_valid(self, form):
        matched_data = MatchedData.objects.get(pk=self.kwargs['pk'])
        generated_data_pk = matched_data.generated.id
        generated = GeneratedData.objects.get(pk=generated_data_pk)
        generated.status = CSV_OUTPUT_COMPLETED
        generated.save()

        matched = form.save(commit=False)
        staff = Staff.objects.get(staff=self.request.user)
        matched.staff = staff
        matched.id = matched_data.id
        matched.generated_id = matched_data.generated_id
        matched.brycen_file = matched_data.brycen_file
        print(f'matched_data:{matched_data}')
        print(f'matched_data.brycen_file:{matched_data.brycen_file}')
        matched.save()

        brycen_file_path = urllib.parse.unquote(matched.brycen_file.path)
        billing_file_path = urllib.parse.unquote(matched.billing_file.path)

        # ファイル命名
        now = datetime.now()
        file_name = 'TXJ_付け合わせ済_' + now.strftime('%Y年%m月%d日%H時%M分%S秒') + '_作成分)' + '.csv'

        output_data = create_csv(brycen_file_path, billing_file_path)

        # ファイルsave
        matched.matched_data_file.save(file_name, ContentFile(output_data))

        return super().form_valid(form)

    def get_success_url(self):
        return reverse('detail_and_create', kwargs={'pk': self.object.pk})


class BillingDataCreateView(LoginRequiredMixin, CreateView):
    form_class = BillingDataFrom
    template_name = 'app/billing_data.html'

    def get_context_data(self, **kwargs):
        matched_data_pk = self.kwargs['pk']
        matched = MatchedData.objects.get(pk=matched_data_pk)
        generated = matched.generated
        formset = BillingDataFromSet(queryset=BillingData.objects.none(), form_kwargs={'customer_id': generated.customer.id})

        context = super().get_context_data(**kwargs)
        context['formset'] = formset
        context['page'] = 'BILLING_DATA'
        context['generated'] = generated
        context['mode'] = 'CREATE'

        return context

    def post(self, request, *args, **kwargs):
        formset = BillingDataFromSet(self.request.POST)
        if formset.is_valid():
            return self.form_valid(formset)

    def form_valid(self, formset):
        matched_data_pk = self.kwargs['pk']
        matched = MatchedData.objects.get(pk=matched_data_pk)
        generated_data_pk = matched.generated.id
        generated = GeneratedData.objects.get(id=generated_data_pk)
        staff = Staff.objects.get(staff=self.request.user)

        if 'save' in self.request.POST:
            self._post_mode = 'SAVE'
            generated.update_date = timezone.now()
            generated.save()

            instances = formset.save(commit=False)
            for instance in instances:
                instance.staff = staff
                instance.matched = matched
                instance.save()

        elif 'save_and_create' in self.request.POST:
            self._post_mode = 'SAVE_AND_CREATE'
            print('save_and_create')
            generated.status = CSV_OUTPUT_COMPLETED
            generated.update_date = timezone.now()
            generated.save()

            instances = formset.save(commit=False)
            for instance in instances:
                instance.staff = staff
                instance.matched = matched
                instance.save()

                if formset.deleted_objects:
                    for delete in formset.deleted_objects:
                        delete.delete()

                billing_data = BillingData.objects.filter(matched_id=matched_data_pk)
                brycen_file_path = urllib.parse.unquote(matched.brycen_file.path)

                # ファイル命名
                now = datetime.now()
                file_name = 'TXJ_付け合わせ済_' + now.strftime('%Y年%m月%d日%H時%M分%S秒') + '_作成分)' + '.csv'

                output_data = create_csv_from_billing_data(billing_data, brycen_file_path)

                # ファイルsave
                matched.matched_data_file.save(file_name, ContentFile(output_data))

        return super().form_valid(formset)

    def get_success_url(self):
        matched_data_pk = self.kwargs['pk']
        if self._post_mode == 'SAVE':
            print('post_mode save')
            matched = MatchedData.objects.get(pk=matched_data_pk)
            billing = BillingData.objects.filter(matched_id=matched_data_pk)
            billing_data_last_row_pk = billing.last().id

            return reverse('billing_data_detail',
                           kwargs={'matched_data_pk': matched.id, 'billing_data_last_row_pk': billing_data_last_row_pk})
        else:
            print('post_mode save_and_create')
            return reverse('detail_and_create', kwargs={'pk': matched_data_pk})


class BillingDataDetailView(LoginRequiredMixin, DetailView):
    model = BillingData
    template_name = 'app/billing_data_detail.html'

    def get_object(self, queryset=None):
        if queryset is None:
            queryset = self.get_queryset()
            new_str = self.kwargs.get('billing_data_last_row_pk') or self.request.GET.get('billing_data_last_row_pk') or None

            queryset = queryset.filter(id=new_str)
            obj = queryset.get()
            return obj

    def get_context_data(self, **kwargs):
        matched_data_pk = self.kwargs['matched_data_pk']
        billing = BillingData.objects.filter(matched_id=matched_data_pk)
        matched = MatchedData.objects.get(id=matched_data_pk)
        billing_data_pk = self.kwargs['billing_data_last_row_pk']
        context = super().get_context_data(**kwargs)

        context['billing'] = billing
        context['matched'] = matched
        context['matched_data_pk'] = matched_data_pk
        context['billing_data_pk'] = billing_data_pk
        context['page'] = 'BILLING_DATA'

        return context


class BillingDataUpdateView(LoginRequiredMixin, UpdateView):
    form_class = BillingDataFrom
    template_name = 'app/billing_data.html'

    def get_queryset(self):
        matched_data_pk = self.kwargs['matched_data_pk']
        queryset = BillingData.objects.filter(matched_id=matched_data_pk)
        return queryset

    def get_object(self, queryset=None):
        if queryset is None:
            queryset = self.get_queryset()
            new_str = self.kwargs.get('billing_data_last_row_pk') or self.request.GET.get('billing_data_last_row_pk') or None

            queryset = queryset.filter(id=new_str)
            obj = queryset.get()
            return obj

    def post(self, request, *args, **kwargs):
        formset = BillingDataFromSet(self.request.POST)
        if formset.is_valid():
            return self.form_valid(formset)

    def form_valid(self, formset):
        matched_data_pk = self.kwargs['matched_data_pk']
        matched = MatchedData.objects.get(pk=matched_data_pk)
        generated_data_pk = matched.generated.id
        generated = GeneratedData.objects.get(id=generated_data_pk)
        staff = Staff.objects.get(staff=self.request.user)
        billing_data = BillingData.objects.filter(matched_id=matched_data_pk)
        for b in billing_data:
            print(b.created_date)

        if 'save' in self.request.POST:
            generated.update_date = timezone.now()
            generated.save()

            instances = formset.save(commit=False)
            for instance in instances:
                instance.staff = staff
                instance.create_date = timezone.now()
                instance.matched = matched
                instance.save()

        if 'save_and_create' in self.request.POST:
            print('save_and_create')
            generated.status = CSV_OUTPUT_COMPLETED
            generated.update_date = timezone.now()
            generated.save()

            instances = formset.save(commit=False)
            for instance in instances:
                instance.staff = staff
                instance.create_date = timezone.now()
                instance.matched = matched
                instance.save()

            if formset.deleted_objects:
                for delete in formset.deleted_objects:
                    delete.delete()

            brycen_file_path = urllib.parse.unquote(matched.brycen_file.path)

            # ファイル命名
            now = datetime.now()
            file_name = 'TXJ_付け合わせ済_' + now.strftime('%Y年%m月%d日%H時%M分%S秒') + '_作成分)' + '.csv'

            output_data = create_csv_from_billing_data(billing_data, brycen_file_path)

            # ファイルsave
            matched.matched_data_file.save(file_name, ContentFile(output_data))

            return HttpResponseRedirect(reverse('detail_and_create', kwargs={'pk': matched_data_pk}))

        return super().form_valid(formset)

    def get_context_data(self, **kwargs):
        matched_data_pk = self.kwargs['matched_data_pk']
        matched = MatchedData.objects.get(pk=matched_data_pk)
        generated = matched.generated
        formset = BillingDataFromSet(queryset=BillingData.objects.filter(matched_id=matched_data_pk), form_kwargs={'customer_id': generated.customer.id})
        billing = BillingData.objects.filter(matched_id=matched_data_pk)
        billing_last = billing.last()
        context = super().get_context_data(**kwargs)
        context['formset'] = formset
        context['page'] = 'BILLING_DATA'
        context['generated'] = generated
        context['billing'] = billing
        context['billing_last'] = billing_last
        context['mode'] = 'EDIT'

        return context

    def get_success_url(self):
        matched_data_pk = self.kwargs['matched_data_pk']
        billing = BillingData.objects.filter(matched_id=matched_data_pk)
        billing_data_pk = billing.last().id
        return reverse('billing_data_detail', kwargs={'matched_data_pk': matched_data_pk, 'billing_data_last_row_pk': billing_data_pk})


# BillingData用突合ロジック
def create_csv_from_billing_data(billing_data_queryset, brycen_file_path):
    BRYCEN_CUSTOMER_NUM = 0
    BRYCEN_STORE_CODE_NUM = 1
    BRYCEN_PT_CODE_NUM = 2
    BRYCEN_UNIT_PRICE_NUM = 4
    BRYCEN_AMOUNT_NUM = 5
    BRYCEN_TOTAL_NUM = 7
    BRYCEN_CONTRACT_START_DATE_NUM = 9
    BRYCEN_CONTRACT_END_DATE_NUM = 13

    WOOD_PALLET_ITEM_NUM = 0
    CONTAINER_REPLACEMENT_ITEM_NUM = 1
    CONTAINER_RENTAL_ITEM_NUM = 2
    STRETCH_FILM_ITEM_NUM = 3
    SLUDGES_ITEM_NUM = 4
    SCRAP_ITEM_NUM = 7
    GENERAL_WASTE_ITEM_NUM = 9
    INDUSTRIAL_WASTE_ITEM_NUM = 10
    MANIFEST_ITEM_NUM = 11
    WASTE_ELEMENT_ITEM_NUM = 12
    WASTE_TIRE_ITEM_NUM = 13
    BASE_TIRE_ITEM_NUM = 14
    WASTE_COOLANT_ITEM_NUM = 15
    WASTE_OIL_ITEM_NUM = 16
    INDUSTRIAL_WASTE_TAX_ITEM_NUM = 17
    WASTE_BATTERY_ITEM_NUM = 18

    KG_UNIT_NUM = 0
    CAR_UNIT_NUM = 1
    ONESET_UNIT_NUM = 2
    MONTHLY_UNIT_NUM = 3
    CUBIC_METER_UNIT_NUM = 4
    TIMES_UNIT_NUM = 5
    CASE_UNIT_NUM = 6
    PEDESTAL_UNIT_NUM = 7
    TIRE_UNIT_NUM = 8
    LITER_UNIT_NUM = 9
    SHEET_UNIT_NUM = 10
    METER_UNIT_NUM = 12
    QUANTITY_UNIT_NUM = 13

    wb = openpyxl.load_workbook(brycen_file_path)
    ws = wb.active
    greatest = ws.max_row

    monthly_fixed_cost_list = []
    spot_cost_list = []
    row_count = 0
    for row in ws.iter_rows(min_col=6, max_row=greatest + 1):
        if row[BRYCEN_CUSTOMER_NUM].value:
            store = row[BRYCEN_STORE_CODE_NUM].value
            store_split = store.split(':')
            store_code = int(store_split[0])
            pt = row[BRYCEN_PT_CODE_NUM].value
            pt_split = pt.split(':')
            pt_code = int(pt_split[0])
            unit_price = row[BRYCEN_UNIT_PRICE_NUM].value
            unit_price = float(unit_price.replace(',', ''))

            amount = row[BRYCEN_AMOUNT_NUM].value
            amount = float(amount.replace(',', ''))

            total = row[BRYCEN_TOTAL_NUM].value
            total = int(total.replace(',', ''))

            contract_start_date = row[BRYCEN_CONTRACT_START_DATE_NUM].value
            contract_end_date = row[BRYCEN_CONTRACT_END_DATE_NUM].value

            # 契約終了日が空欄だったら
            if contract_end_date is None:
                row_count += 1
                monthly_fixed_cost = [contract_start_date, store_code, pt_code, unit_price, amount, total]
                monthly_fixed_cost_list.append(monthly_fixed_cost)

            # 契約終了日に日付が入っていたら
            elif contract_end_date:
                row_count += 1
                spot_cost = [contract_start_date, store_code, pt_code, unit_price, amount, total]
                spot_cost_list.append(spot_cost)

            else:
                other_list = [contract_start_date, contract_end_date, store_code, pt_code, unit_price, amount, total]
                print(f'other_list:{other_list}')

    items_dict = {
        WOOD_PALLET_ITEM_NUM: '木パレット',
        CONTAINER_REPLACEMENT_ITEM_NUM: 'コンテナ交換',
        CONTAINER_RENTAL_ITEM_NUM: 'コンテナレンタル代',
        STRETCH_FILM_ITEM_NUM: 'ストレッチフィルム',
        SLUDGES_ITEM_NUM: '汚泥',
        SCRAP_ITEM_NUM: 'スクラップ類',
        GENERAL_WASTE_ITEM_NUM: '一般廃棄物',
        INDUSTRIAL_WASTE_ITEM_NUM: '産業廃棄物',
        MANIFEST_ITEM_NUM: 'マニフェスト',
        WASTE_ELEMENT_ITEM_NUM: '廃エレメント',
        WASTE_TIRE_ITEM_NUM: '廃タイヤ',
        BASE_TIRE_ITEM_NUM: '台タイヤ',
        WASTE_COOLANT_ITEM_NUM: '廃クーラント',
        WASTE_OIL_ITEM_NUM: '廃油',
        INDUSTRIAL_WASTE_TAX_ITEM_NUM: '産廃税',
        WASTE_BATTERY_ITEM_NUM: '廃バッテリー'
    }

    units_dict = {
        KG_UNIT_NUM: 'kg',
        CAR_UNIT_NUM: '車',
        ONESET_UNIT_NUM: '式',
        MONTHLY_UNIT_NUM: '月額',
        CUBIC_METER_UNIT_NUM: '立米',
        TIMES_UNIT_NUM: '回',
        CASE_UNIT_NUM: 'ケース',
        PEDESTAL_UNIT_NUM: '台',
        TIRE_UNIT_NUM: '本',
        LITER_UNIT_NUM: 'リットル',
        SHEET_UNIT_NUM: '枚',
        METER_UNIT_NUM: 'メートル',
        QUANTITY_UNIT_NUM: '個'
    }
    billing_data = []
    writer_data = []
    for b in billing_data_queryset:
        date = str(b.billing_date)
        date_split = date.split('-')
        year = date_split[0]
        month = date_split[1]
        day = date_split[2]
        billing_date = f'{year}/{month}/{day}'
        agent_name = b.agent.name
        agent_code = int(b.agent.code)
        customer_code = int(b.place.customer.code)
        place_name = b.place.name
        place_code = int(b.place.code)
        product_name = b.product.name
        product_code = int(b.product.code)
        item_num = b.item
        for k, v in items_dict.items():
            if item_num == k:
                item_name = v
        amount = b.amount
        unit_num = b.unit
        for k, v in units_dict.items():
            if unit_num == k:
                unit = v
        unit_price = b.unit_price
        total = b.total

        billing_list = [billing_date, place_code, agent_code, unit_price, amount, total]
        billing_data.append(billing_list)

        all_billing_data = [customer_code, place_code, place_name, billing_date, agent_code, agent_name,
                            product_code, item_name, unit_price, amount, unit, total]

        PLACE_CODE_INDEX = 1
        BILLING_DATE_INDEX = 3
        AGENT_CODE_INDEX = 4
        UNIT_PRICE_INDEX = 8
        AMONT_INDEX = 9
        TOTAL_INDEX = 11

        compare = [all_billing_data[BILLING_DATE_INDEX], all_billing_data[PLACE_CODE_INDEX], all_billing_data[AGENT_CODE_INDEX],
                   all_billing_data[UNIT_PRICE_INDEX], all_billing_data[AMONT_INDEX], all_billing_data[TOTAL_INDEX]]

        # 月極費用と請求データの[事業所コード、業者コード、単価、数量、合計金額]が
        # 完全一致した行をbilling_dataからremove
        for b in billing_data:
            for m in monthly_fixed_cost_list:
                if m[1:] == b[1:]:
                    print(f'monthly_cost_list billing_list match row:{b}')
                    billing_data.remove(b)

            # スポット費用と請求データの[契約開始日（請求日）、事業所コード、業者コード、単価、数量、合計金額]が
            # 完全一致した行をbilling_dataからremove
            for s in spot_cost_list:
                if b == s:
                    print(f'spot_cost billing_data match row:{s}')
                    billing_data.remove(s)

    # billing_dataのままだとcustomer_code等の情報が欠けている。欠けている情報を補う処理。
    # all_billing_data(大元の請求データ)を元に作成したcompare(billing_data比較用のリスト)が
    # billing_dataに含まれる場合は、writer_data.append
    if compare in billing_data:
        writer_data.append(all_billing_data)
        # print(writer_data)

    output = io.StringIO()
    header = ['取引先名', '支店番号', '支店名', '日付', '業者番号', '業者名', '商品コード', '品目', '単価', '数量', '単位', '合計金額', '備考']
    writer = csv.writer(output, quoting=csv.QUOTE_NONNUMERIC)
    writer.writerow(header)
    writer.writerows(writer_data)
    output_data = output.getvalue() # .encode('cp932')

    return output_data


class MatchedDataDetailAndVisuallyMatchedDataCreateView(LoginRequiredMixin, CreateView):
    model = MatchedData
    form_class = VisuallyMatchedDataCreateForm
    template_name = 'app/detail_and_create.html'

    def get_context_data(self, **kwargs):
        matched_data_pk = self.kwargs['pk']
        matched = MatchedData.objects.get(pk=matched_data_pk)
        billing = BillingData.objects.filter(matched_id=matched_data_pk)
        context = super().get_context_data(**kwargs)
        if billing.exists():
            billing_data_last_row_pk = billing.last().pk
            context['billing_data_last_row_pk'] = billing_data_last_row_pk
        context['matched'] = matched
        context['status'] = CSV_OUTPUT_COMPLETED
        context['billing'] = billing

        return context

    def form_valid(self, form):
        visually_matched = form.save(commit=False)
        staff = Staff.objects.get(staff=self.request.user)
        visually_matched.staff = staff
        matched_data_pk = self.kwargs['pk']
        matched = MatchedData.objects.get(pk=matched_data_pk)
        visually_matched.matched_id = matched.id
        visually_matched.save()

        generated_data_pk = visually_matched.matched.generated.id
        generated = GeneratedData.objects.get(pk=generated_data_pk)
        generated.status = VISUALLY_CONFIRMED
        generated.save()

        return super().form_valid(form)

    def get_success_url(self):
        return reverse('import_data', kwargs={'pk': self.object.pk})


# 突合スクリプト
def create_csv(f, f2):
    BRYCEN_STORE_CODE_NUM = 1
    BRYCEN_PT_CODE_NUM = 2
    BRYCEN_UNIT_PRICE_NUM = 4
    BRYCEN_AMOUNT_NUM = 5
    BRYCEN_TOTAL_NUM = 7
    BRYCEN_CONTRACT_START_DATE_NUM = 9
    BRYCEN_CONTRACT_END_DATE_NUM = 13

    BILLING_STORE_CODE_NUM = 49
    BILLING_STORE_NAM_NUM = 50
    BILLING_DAY_NUM = 39
    BILLING_PT_CODE_NUM = 2
    BILLING_PT_NAM_NUM = 3
    BILLING_ITEM_CODE_NUM = 41
    BILLING_ITEM_NAM_NUM = 42
    BILLING_UNIT_PRICE_NUM = 43
    BILLING_UNIT_NUM = 44
    BILLING_AMOUNT_NUM = 45
    BILLING_TOTAL_NUM = 46
    BILLING_REMARK_NUM = 51
    BILLING_SUBJECT_NUM = 8

    BRYCEN_STORE_CODE_LIST_INDEX = 0
    BRYCEN_PT_CODE_LIST_INDEX = 1
    BRYCEN_TOTAL_LIST_INDEX = 4

    STORE_CODE_LIST_INDEX = 0
    PT_CODE_LIST_INDEX = 1
    TOTAL_LIST_INDEX = 4

    zenTOLL = 'トール'
    hanTOLL = 'ﾄｰﾙ'
    otherTOLL = '九州産交'

    print('##########csv_create start############')
    wb = openpyxl.load_workbook(f)
    ws = wb.active
    greatest = ws.max_row

    # 契約データの必要情報を取得する
    monthly_fixed_cost_list = []
    spot_cost_list = []
    row_count = 0
    for row in ws.iter_rows(min_col=6, max_row=greatest + 1):
        if row[0].value:
            store = row[BRYCEN_STORE_CODE_NUM].value
            store_split = store.split(':')
            store_code = int(store_split[0])

            pt = row[BRYCEN_PT_CODE_NUM].value
            pt_split = pt.split(':')
            pt_code = int(pt_split[0])
            unit_price = row[BRYCEN_UNIT_PRICE_NUM].value
            unit_price = float(unit_price.replace(',', ''))

            amount = row[BRYCEN_AMOUNT_NUM].value
            amount = float(amount.replace(',', ''))

            total = row[BRYCEN_TOTAL_NUM].value
            total = int(total.replace(',', ''))


            contract_start_date = row[BRYCEN_CONTRACT_START_DATE_NUM].value
            contract_end_date = row[BRYCEN_CONTRACT_END_DATE_NUM].value

            # 契約終了日が空欄だったら
            if contract_end_date is None:
                row_count += 1
                monthly_fixed_cost = [contract_start_date, store_code, pt_code, unit_price, amount, total]
                monthly_fixed_cost_list.append(monthly_fixed_cost)

            # 契約終了日に日付が入っていたら
            elif contract_end_date:
                row_count += 1
                spot_cost = [contract_start_date, store_code, pt_code, unit_price, amount, total]
                spot_cost_list.append(spot_cost)

            else:
                other_list = [contract_start_date, contract_end_date, store_code, pt_code, unit_price, amount, total]
                print(f'other_list:{other_list}')
                #break

            # 現行スクリプト
            # brycen_list = [store_code, pt_code, unit_price, amount, total]
            # brycen_data.append(brycen_list)

    file_b = open(f2, mode='rb')
    file_b_read = file_b.read()
    binary = chardet.detect(file_b_read)
    print('billing_file binary:',binary)
    encoding = binary['encoding']
    if binary['encoding'] == 'CP932':
        print(f'encoding is {encoding}')
        file = open(f2, encoding="CP932", errors='replace')

    elif binary['encoding'] == 'SHIFT_JIS':
        print(f'encoding is {encoding}')
        file = open(f2, encoding='SHIFT_JIS', errors='replace')
    else:
        print(f'encoding is {encoding}')
        file = open(f2, encoding="utf-8", errors='replace')
    reader = csv.reader(file)
    data = list(reader)
    data.remove(data[0])

    # 電子請求データの必要情報を取得する　
    all_store_code = {'2493-1': '千葉支店', '2493-2': '板橋支店', '2493-3': '鹿島支店', '2493-4': '東京北支店', '2493-5': '東京支店',
                      '2493-6': '千葉南支店',
                      '2493-7': '土浦支店', '2493-8': '大宮支店', '2493-9': '熊谷支店', '2493-10': '市原支店', '2493-11': '東京中央支店',
                      '2493-12': '横浜支店',
                      '2493-13': '港北支店', '2493-14': '戸塚支店', '2493-15': '埼玉支店', '2493-17': '厚木支店', '2493-18': '青梅支店',
                      '2493-19': '平塚支店', '2493-20': '静岡支店', '2493-21': '三島支店',
                      '2493-22': '富士支店', '2493-23': '浜松支店', '2493-24': '掛川支店', '2493-27': '小牧支店', '2493-28': '岐阜支店',
                      '2493-29': '中部支店', '2493-30': '四日市支店',
                      '2493-31': '滋賀支店', '2493-33': '近江八幡支店', '2493-34': '京都支店', '2493-35': '大阪支店',
                      '2493-36': '南港支店', '2493-37': '貝塚支店', '2493-38': '奈良支店', '2493-39': '東大阪支店',
                      '2493-40': '松原支店', '2493-42': '和歌山支店', '2493-43': '淡路支店', '2493-44': '西神戸支店',
                      '2493-45': '三木小野支店', '2493-46': '加古川支店', '2493-47': '西脇支店', '2493-48': '福知山支店',
                      '2493-49': '姫路支店', '2493-50': '岡山支店', '2493-51': '福山支店', '2493-52': '広島支店', '2493-53': '徳山支店',
                      '2493-54': '山口支店', '2493-55': '高松支店', '2493-56': '新居浜支店', '2493-57': '松山支店',
                      '2493-58': '徳島支店', '2493-59': '丸亀支店', '2493-60': '福岡支店', '2493-61': '大牟田支店',
                      '2493-62': '中津支店', '2493-63': '北九州支店', '2493-100': 'トールエクスプレスジャパン　本社', '2493-108': '羽生支店',
                      '2493-111': '尼崎支店', '2493-112': '中九州支店', '2493-113': '長崎支店', '2493-114': '大分支店',
                      '2493-115': '熊本支店', '2493-116': '佐世保支店', '2493-117': '鹿児島支店', '2493-118': '佐賀支店',
                      '2493-119': '天草支店', '2493-120': '八代支店', '2493-121': '人吉支店', '2493-122': '宮崎支店',
                      '2493-123': '日向支店', '2493-124': '水俣支店', '2493-125': '姶良支店', '2493-126': '川内支店',
                      '2493-130': '安城支店', '2493-131': '名古屋支店', '2505-1': '千葉工場', '2505-2': '厚木工場', '2505-3': '浜松工場',
                      '2505-4': '中部工場', '2505-5': '四日市工場', '2505-6': '大阪工場', '2505-7': '東大阪工場', '2505-8': '西脇工場',
                      '2505-11': '福山工場', '2505-13': '高松工場', '2505-14': '九州工場', '2505-15': '福岡工場', '2505-17': '中九州工場',
                      '2505-18': '熊本工場', '2505-19': '八代工場', '2508-9': '長崎センター', '2508-18': '宮崎センター',
                      '2508-19': '栗野食品センター',
                      '2508-20': '熊本センター', '2508-22': '福岡センター', '2508-23': '九州産交運輸　本社', '2508-29': '鹿児島センター',
                      '2508-30': '佐賀事業所', '2508-32': '熊本コンテナセンター', '2508-33': '八代コンテナ事業所',
                      '2508-35': '福岡コンテナ事業所',
                      '2508-36': '北九州コンテナ事業所', '2508-37': '中九州センター'}

    billing_data = []
    writer_data = []
    pass_data = []

    for row in data:
        unit_price = row[BILLING_UNIT_PRICE_NUM]
        if not unit_price:
            unit_price = ''
        else:
            unit_price = float(unit_price)

        amount = row[BILLING_AMOUNT_NUM]
        if not amount:
            amount = ''
        else:
            amount = float(amount)

        total = row[BILLING_TOTAL_NUM]
        if not total:
            total = ''

        else:
            total = int(total)

        store_code = row[BILLING_STORE_CODE_NUM]
        store_nam = row[BILLING_STORE_NAM_NUM]
        billing_date = row[BILLING_DAY_NUM]
        pt_code = int(row[BILLING_PT_CODE_NUM])
        pt_nam = row[BILLING_PT_NAM_NUM]
        item_nam = row[BILLING_ITEM_NAM_NUM]
        item_code = row[BILLING_ITEM_CODE_NUM]
        unit = row[BILLING_UNIT_NUM]
        remark = row[BILLING_REMARK_NUM]

        # トール以外の行をなるべく外す処理
        # 1.備考欄に0がある 2.部門名or備考欄に「トールorﾄｰﾙor九州産交」が含まれる 3.部門コードor備考欄に取引先コードが含まれる
        # 4.部門名に値がある　1~4に当てはまればpassされない
        if ('0' not in remark) and (zenTOLL not in store_nam and zenTOLL not in item_nam and zenTOLL not in remark and \
            hanTOLL not in store_nam and hanTOLL not in item_nam and hanTOLL not in remark and
            otherTOLL not in store_nam and otherTOLL not in item_nam and otherTOLL not in remark ) and \
                ('2493' not in store_code and '2493' not in remark and not store_nam and
                 '2505' not in store_code and '2505' not in remark and '2508' not in store_code and '2508' not in remark):
            # 1~4に当てはまらなくても下記pt_codeの請求行はpassされない
            if int(pt_code) not in [1023, 1100, 2857, 2877, 3350, 3361, 3376, 3353, 3370, 3381, 3422, 3443, 3742, 3761, 5145, 5172]:
                pass_data.append([store_code, store_nam, item_nam, remark, pt_code, pt_nam])
                continue

        # store_codeが無ければ(部門コード列空欄)、dictのkey(取引先コード-事業所コード)をいれる
        if not store_code:
            for key, value in all_store_code.items():
                # store_codeがない場合、store_nam or remark or item_numに事業所名が一致するものがないか確認
                if (store_nam is not None and value == store_nam) or (remark is not None and value == remark) or (
                        item_nam is not None and value == item_nam):
                    store_code = str(key)
                    store_nam = value
                    break

        # store_codeはある前提。if not store_code:〜store_code = str(key)の間で条件一致していれば、store_codeは空にならない
        # dictに一致しないということは、そもそもトール関連の請求ではない可能性が高い
        if store_code:
            if not store_nam:
                for key, value in all_store_code.items():
                    if store_code in key:
                        store_nam = value
                        break

            store_split = store_code.split('-')
            customer_mach = re.search(r'(\d{4})', str(store_split[0]))

            if not customer_mach:
                #customer_code = store_split[0]
                #store_code = int(store_split[1])
                # print(store_nam)
                if store_nam:
                    for key, value in all_store_code.items():
                        if store_nam == value:
                            key_split = key.split('-')
                            customer_code = key_split[0]
                            store_code = key_split[1]
                else:
                    store_nam = '未入力'

            else:
                customer_code = store_split[0]
                store_code = int(store_split[1])


        # この段階で、さらにstore_codeがなければ、どうやって処理を続けるか考える？
        # 更に、store_codeもstore_namもない行に対しての処理はココ！
        else:
            store_code = '未入力'
            customer_code = '未入力'


        #if int(pt_code) == 3422:
            #if 'プラ' in remark or 'ﾌﾟﾗ' in remark or 'プラ' in item_nam or 'ﾌﾟﾗ' in item_nam:
                #item_nam = 'プラパレット'
            #else:
                #item_nam = '木パレット'
        if int(pt_code) == 3742:
            if store_code == '未入力':
                customer_code = '2493'
                item_nam = '木パレット'
                if '若松' in row[BILLING_SUBJECT_NUM]:
                    store_code = '63'
                    store_nam = '北九州支店'
                else:
                    store_code = '34'
                    store_nam = '京都支店'

        billing_list = [billing_date, store_code, pt_code, unit_price, amount, total]
        billing_data.append(billing_list)

        all_billing_data = [customer_code, store_code, store_nam, billing_date, pt_code, pt_nam, item_code, item_nam,
                            unit_price, amount, unit, total, remark]

        STORE_CODE_INDEX = 1
        BILLING_DATE_INDEX = 3
        PT_CODE_INDEX = 4
        UNIT_PRICE_INDEX = 8
        AMONT_INDEX = 9
        TOTAL_INDEX = 11

        # compare = [all_billing_data[1], all_billing_data[4], all_billing_data[8], all_billing_data[9],
                   #all_billing_data[11]]
        compare = [all_billing_data[BILLING_DATE_INDEX], all_billing_data[STORE_CODE_INDEX], all_billing_data[PT_CODE_INDEX],
                   all_billing_data[UNIT_PRICE_INDEX], all_billing_data[AMONT_INDEX], all_billing_data[TOTAL_INDEX]]

        # billing_list = [billing_date, store_code, pt_code, unit_price, amount, total]
        # print(billing_list)

        # for b in brycen_data:
        for b in billing_data:
            for m in monthly_fixed_cost_list:
                #print(f'billing_data:{b[1:]}')
                #print(f'billing_data:{b[1:]} monthly:{m[1:]}')

                # 月極費用と請求データの[事業所コード、業者コード、単価、数量、合計金額]が
                # 完全一致した行をbilling_dataからremove
                if b[1:] == m[1:]:
                    print(f'monthly_cost billing_data match row:{b}')
                    billing_data.remove(b)

            # スポット費用と請求データの[契約開始日、事業所コード、業者コード、単価、数量、合計金額]が
            # 完全一致した行をbilling_dataからremove
            for s in spot_cost_list:
                if b == s:
                    print(f'spot_cost billing_data match row:{s}')
                    billing_data.remove(s)
                    # print(f'remove spot:{billing_data}')

            # 現行スクリプト
            # if billing_list == b:
                # print('remove{}:'.format(b))
                # print('mach:',billing_list)
                # billing_data.remove(b)
                # print('remove')
                #billing_data_day.remove(b)

        # billing_dataのままだとcustomer_code等の情報が欠けている。欠けている情報を補う処理。
        # all_billing_data(大元の請求データ)を元に作成したcompare(billing_data比較用のリスト)が
        # billing_dataに含まれる場合は、writer_data.append
        if compare in billing_data:
            writer_data.append(all_billing_data)
            # print(f'all_billing_data:{all_billing_data}')

        # 現行スクリプト
        # brycen_dataとbilling_dataが不一致だった場合のみ
        # writer_dataにappend
        #if compare in billing_data:
            # writer_data.append(all_billing_data)
            # print(all_billing_data)

    output = io.StringIO()
    header = ['取引先名', '支店番号', '支店名', '日付', '業者番号', '業者名', '商品コード', '品目', '単価', '数量', '単位', '合計金額', '備考']
    writer = csv.writer(output,  quoting=csv.QUOTE_NONNUMERIC)
    writer.writerow(header)
    writer.writerows(writer_data)
    output_data = output.getvalue().encode('cp932')

    # TXJ以外の請求行（突合済ファイルに反映されない請求データ）をprintする。
    for pass_item in pass_data:
        #print(
        #'============= Pass \n store: {} {} \n item_remark : {}_{} \n pt    :{} {} ============'.format(
        #pass_list = ['============= Pass \n store: {} {} \n item_remark : {}_{} \n pt    :{} {} ============'.format(
        pass_list = [f'store:{pass_item[0]} {pass_item[1]}, item:{pass_item[2]}, remark:{pass_item[3]}, '
                     f'pt:{pass_item[4]} {pass_item[5]}']

    return output_data


class ImportDataCreateView(LoginRequiredMixin, CreateView):
    model = ImportData
    form_class = ImportDataCreateForm
    template_name = 'app/import_data.html'

    def get_context_data(self, **kwargs):
        visually_matched_data_pk = self.kwargs['pk']
        visually_matched = VisuallyMatchedData.objects.get(pk=visually_matched_data_pk)
        matched_data_pk = visually_matched.matched_id
        matched = MatchedData.objects.get(pk=matched_data_pk)

        context = super().get_context_data(**kwargs)
        context['visually_matched'] = visually_matched
        context['matched'] = matched
        context['status'] = VISUALLY_CONFIRMED
        return context

    def get_form_kwargs(self, *args, **kwargs):
        form_kwargs = super().get_form_kwargs(*args, **kwargs)
        form_kwargs.update({'method': self.request.method})
        if self.request.method == 'POST':
            if 'upload_and_create' in self.request.POST:
                form_kwargs.update({'upload_and_create': self.request.POST.get('upload_and_create', None) is not None})

        return form_kwargs

    @transaction.atomic
    def form_valid(self, form):
        # ファイル命名
        now = datetime.now()
        file_name = 'TXJ_import_data_' + now.strftime('%Y年%m月%d日%H時%M分%S秒') + '_作成分' + '.xlsx'

        if 'upload_and_create' in self.request.POST:
            import_data = form.save(commit=False)
            staff = Staff.objects.get(staff=self.request.user)
            import_data.staff = staff
            visually_matched_data_pk = self.kwargs['pk']
            visually_matched = VisuallyMatchedData.objects.get(pk=visually_matched_data_pk)
            import_data.visually_matched_id = visually_matched.id
            import_data.save()

            if import_data.visually_matched_file:
                visually_matched_file_path = urllib.parse.unquote(import_data.visually_matched_file.path)
                output_data = import_data_create(visually_matched_file_path)
                import_data.import_data_file.save(file_name, ContentFile(output_data))

        elif 'create' in self.request.POST:
            import_data = form.save(commit=False)
            staff = Staff.objects.get(staff=self.request.user)
            import_data.staff = staff
            visually_matched_data_pk = self.kwargs['pk']
            visually_matched = VisuallyMatchedData.objects.get(pk=visually_matched_data_pk)
            import_data.visually_matched_id = visually_matched.id
            matched_file_path = urllib.parse.unquote(visually_matched.matched.matched_data_file.path)

            output_data = import_data_create(matched_file_path)
            import_data.import_data_file.save(file_name, ContentFile(output_data))
            import_data.save()

        generated_data_pk = import_data.visually_matched.matched.generated.pk
        generated = GeneratedData.objects.get(pk=generated_data_pk)
        generated.status = GeneratedData.IMPORT_DATA_OUTPUT_COMPLETED
        generated.save()


        return super().form_valid(form)

    def get_success_url(self):
        return reverse('import_data_detail', kwargs={'pk': self.object.pk})


# インポートデータ作成ロジック
def import_data_create(f):
    STORE_CODE_LIST_INDEX = 0
    DAY_LIST_INDEX = 1
    PT_CODE_LIST_INDEX = 2
    UNIT_PRICE_LIST_INDEX = 3
    AMOUNT_LIST_INDEX = 4
    UNIT_LIST_INDEX = 5
    TOTAL_LIST_INDEX = 6
    REMARK_LIST_INDEX = 7
    PRODUCT_CODE_LIST_INDEX = 8
    PRODUCT_NAM_LIST_INDEX = 9
    ITEM_CODE_LIST_INDEX = 10
    UNIT_CODE_LIST_INDEX = 11

    STORE_CODE_COL_NUM = 1
    DAY_COL_NUM = 3
    PT_CODE_COL_NUM = 4
    PRODUCT_COL_NUM = 6
    ITEM_COL_NUM = 7
    UNIT_PRICE_COL_NUM = 8
    AMOUNT_COL_NUM = 9
    UNIT_COL_NUM = 10
    TOTAL_COL_NUM = 11

    print('#########import_data_create start#########')
    # 単位がtできた請求データをkgに変換する関数
    def kg_conversion():
        kg = math.ceil(amount * 1000) # <class 'int'>
        l[AMOUNT_LIST_INDEX] = kg

        kg_unit_price = int(total / kg)
        l[UNIT_PRICE_LIST_INDEX] = kg_unit_price

        l[UNIT_LIST_INDEX] = 'kg'
        l[UNIT_CODE_LIST_INDEX] = '0'

    # 数量を備考欄に反映させ、数量を1式に変換する関数
    def kg_remake():
        l[UNIT_PRICE_LIST_INDEX] = total
        l[AMOUNT_LIST_INDEX] = 1
        l[REMARK_LIST_INDEX] = str(unit_price) + '円' + '×' + '{:,}'.format(int(amount)) + 'kg'
        l[REMARK_LIST_INDEX] = f'{unit_price}円 x {int(amount):,}kg'
        l[UNIT_CODE_LIST_INDEX] = '2'

    file_b = open(f, mode='rb')
    file_b_read = file_b.read()
    binary = chardet.detect(file_b_read)
    print('file binary', binary)
    if binary['encoding'] == 'CP932':
        print('encoding is CP932')
        file = open(f, encoding="CP932", errors='replace')

    elif binary['encoding'] == 'SHIFT_JIS':
        print('encoding is SHIFT_JIS')
        file = open(f, encoding="SHIFT_JIS", errors='replace')
    else:
        print('encoding is utf-8')
        file = open(f, encoding="utf-8", errors='replace')

    #file = open(f, encoding="utf8", errors='replace')  # csvファイルを読み込んだ内容をfileに入れる
    reader = csv.reader(file)  # csvファイルを開いて、開いた内容をreaderに入れる　
    data = list(reader)  # 開いた内容をリストで取得して、dataに入れる

    #  pt_codeが10101の単価15だった場合の行数を取得
    n_row_count = 1
    for n in data:
        if n[PT_CODE_COL_NUM] == '10101' and n[UNIT_PRICE_COL_NUM] == '15':
            n_row_count += 1

    xlsx_path = os.path.join(settings.MEDIA_ROOT, 'import_data_format.xlsx')
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active

    current_dir = os.getcwd()

    # csvファイルの1行目（項目）を削除
    data.remove(data[0])

    i = 6  # 実際にExcelへ書き出す行数（6行目から書き出し開始）
    row_count = 0  # Excelに書き出された行数をカウント

    # csvから必要な値を取得
    for row in data:
        store_code = row[STORE_CODE_COL_NUM]
        if store_code == '未入力':
            continue
        store_code = str(store_code.rjust(5, '0'))

        match = re.search(r'(\d{4})/(\d+)/(\d+)', str(row[DAY_COL_NUM]))
        year = int(match.group(1))
        month = int(match.group(2))
        date = int(match.group(3))
        conv_date = datetime(year, month, date) - datetime(1899, 12, 30)
        day = int(conv_date.days)
        pt_code = str(row[PT_CODE_COL_NUM].rjust(5, '0'))

        amount = row[AMOUNT_COL_NUM]
        if amount:
            amount = float(amount)

        unit = row[UNIT_COL_NUM]

        total = int(row[TOTAL_COL_NUM])

        product_code = row[PRODUCT_COL_NUM]
        if product_code:
            product_code = int(product_code)

        item = row[ITEM_COL_NUM]

        product_name = None
        item_code = None

        # PT：10101の単価15.0の行はlに入れない
        if pt_code == '10101':
            if day == day:
                if int(float(row[UNIT_PRICE_COL_NUM])) == 15:
                    continue

        # 単価はfloatで取得
        if row[UNIT_PRICE_COL_NUM]:
            unit_price = float(row[UNIT_PRICE_COL_NUM])

        # 単価の入力が無かったら、空文字のまま取得
        else:
            if not row[UNIT_PRICE_COL_NUM]:
                unit_price = ''

        # TODO:商品コードを判定し、商品名、商品コード、品目コードを設定する
        if product_code == 4:
            product_code = '0004'
            product_name = '産廃税'
            item_code = '17'

        elif product_code == 110:
            product_code = '0110'
            product_name = '機密文書処理'
            # item_code = '対象コードがありません：品目を確認してください'
            item_code = ''

        elif product_code == 123:
            product_code = '0123'
            product_name = '産業廃棄物（スポット  廃タイヤ）収集運搬処分費'
            item_code = '13'

        elif product_code == 202:
            product_code = '0202'
            product_name = 'マニフェスト伝票代'
            item_code = '11'

        elif product_code == 404:
            product_code = '0404'
            product_name = 'コンテナレンタル費用'
            item_code = '2'

        # product_codeが5103の場合のみitem（品目列）が、
        # コンテナ交換or産業廃棄物かで、item_codeが異なる
        # ただし、上記item以外の場合は全て産業廃棄物のitem_codeになる
        elif product_code == 5103:
            product_code = '5103'
            product_name = '産業廃棄物収集運搬処分費'

            if item == 'コンテナ交換':
                item_code = '1'

            elif item == '産業廃棄物':
                item_code = '10'

            else:
                item_code = '10'

        elif product_code == 5105:
            product_code = '5105'
            product_name = '産業廃棄物（木パレット）収集運搬処分費'
            item_code = '0'

        elif product_code == 5106:
            product_code = '5106'
            product_name = '産業廃棄物（プラパレット）収集運搬処分費'
            item_code = '10'

        elif product_code == 5107:
            product_code = '5107'
            product_name = '産業廃棄物（ストレッチフィルム）収集運搬処分費'
            item_code = '3'

        elif product_code == 9004 and unit_price <= 0:
            product_code = '9004'
            product_name = '産業廃棄物（スポット スクラップ類）収集運搬処分費（買取）'
            item_code = '7'

        elif product_code == 9008 and unit_price <= 0:
            product_code = '9008'
            product_name = 'リサイクル物運搬費（買取）'
            item_code = '18'

        elif product_code == 9012 and unit_price <= 0:
            product_code = '9012'
            product_name = '資源（ストレッチフィルム）収集運搬処分費（買取）'
            item_code = '3'

        elif product_code == 9013 and unit_price <= 0:
            product_code = '9013'
            product_name = '資源（廃油）収集運搬処分費（買取）'
            item_code = '16'

        elif product_code == 9014 and unit_price <= 0:
            product_code = '9014'
            product_name = '資源（台タイヤ）収集運搬処分費（買取）'
            item_code = '14'

        elif product_code == 9015 and unit_price <= 0:
            product_code = '9015'
            product_name = '廃油収集運搬処分費（買取）'
            item_code = '16'

        elif product_code == 9016 and unit_price <= 0:
            product_code = '9016'
            product_name = '資源物収集運搬処分費（買取）'
            # item_code = '対象コードがありません：品目を確認してください'
            item_code = ''

        else:
            pass
            # print(product_code)

            # TODO:明細項目を判定し、商品名、商品コード、品目コードを設定する
            if item == '産廃税':
                product_code = '0004'
                product_name = '産廃税'
                item_code = '17'

            elif item == '機密文書':
                product_code = '0110'
                product_name = '機密文書処理'
                # item_code = '対象コードがありません：品目を確認してください'
                item_code = ''

            elif item == '廃タイヤ':
                product_code = '0123'
                product_name = '産業廃棄物（スポット  廃タイヤ）収集運搬処分費'
                item_code = '13'

            elif item == 'マニフェスト':
                product_code = '0202'
                product_name = 'マニフェスト伝票代'
                item_code = '2'

            elif item == 'コンテナ交換':
                product_code = '5103'
                product_name = '産業廃棄物収集運搬処分費'
                item_code = '1'

            elif item == '産業廃棄物':
                product_code = '5103'
                product_name = '産業廃棄物収集運搬処分費'
                item_code = '10'

            elif item == '木パレット':
                product_code = '5105'
                product_name = '産業廃棄物（木パレット）収集運搬処分費'
                item_code = '0'

            elif item == 'プラパレット':
                product_code = '5106'
                product_name = '産業廃棄物（プラパレット）収集運搬処分費'
                item_code = '10'

            elif item == 'ストレッチフィルム':
                product_code = '5107'
                product_name = '産業廃棄物（ストレッチフィルム）収集運搬処分費'
                item_code = '3'

            elif item == 'スクラップ類　買取' and unit_price <= 0:
                product_code = '9004'
                product_name = '産業廃棄物（スポット スクラップ類）収集運搬処分費（買取）'
                item_code = '7'

            elif item == '廃バッテリー　買取' and unit_price <= 0:
                product_code = '9008'
                product_name = 'リサイクル物運搬費（買取）'
                item_code = '18'


            elif item == 'ストレッチフィルム　買取' and unit_price <= 0:
                product_code = '9012'
                product_name = '資源（ストレッチフィルム）収集運搬処分費（買取）'
                item_code = '3'


            elif item == '廃油　買取' and unit_price <= 0:
                product_code = '9013'
                product_name = '資源（廃油）収集運搬処分費（買取）'
                item_code = '16'


            elif item == '台タイヤ　買取' and unit_price <= 0:
                product_code = '9014'
                product_name = '資源（台タイヤ）収集運搬処分費（買取）'
                item_code = '14'


            elif item == '廃油　買取' and unit_price <= 0:
                product_code = '9015'
                product_name = '廃油収集運搬処分費（買取）'
                item_code = '16'


            elif item == '資源物　買取' and unit_price <= 0:
                product_code = '9016'
                product_name = '資源物収集運搬処分費（買取）'
                # item_code = '対象コードがありません：品目を確認してください'
                item_code = ''

            else:
                pass
                # print(item)

        # TODO:単位列を判定し、請求書記載単位コードを設定する
        if unit == 'ｋｇ':
            unit_code = '0'

        elif unit == 'Ｋｇ':
            unit_code = '0'

        elif unit == 'kg':
            unit_code = '0'

        elif unit == 'Kg':
            unit_code = '0'

        elif unit == '㎏':
            unit_code = '0'

        elif unit == '車':
            unit_code = '1'

        elif unit == '式':
            unit_code = '2'

        elif unit == '立法メートル':
            unit_code = '4'

        elif unit == '立米':
            unit_code = '4'

        elif unit == 'm3':
            unit_code = '4'

        elif unit == '回':
            unit_code = '5'

        elif unit == '台':
            unit_code = '7'

        elif unit == '枚':
            unit_code = '10'

        elif unit == '本':
            unit_code = '8'

        elif unit == 'リットル':
            unit_code = '9'

        elif unit == 'm':
            unit_code = '12'

        elif unit == 'ｍ':
            unit_code = '12'

        elif unit == 'メートル':
            unit_code = '12'

        elif unit == '個':
            unit_code = '13'

        else:
            # unit_code = '単位を確認してください'
            unit_code = ''
            print(unit)
        l = [store_code, day, pt_code, unit_price, amount, unit, total, '', product_code, product_name, item_code,
             unit_code]
        print(l)

        # TODO:イレギュラーPTを判定　→　特殊計算ロジック
        if pt_code == '03422':
            if unit_price == 14.5:
                kg_remake()

            elif unit_price == 4.5:
                kg_remake()

            elif unit_price == 12:
                l[UNIT_PRICE_LIST_INDEX] = total
                l[AMOUNT_LIST_INDEX] = 1
                l[REMARK_LIST_INDEX] = str(int(unit_price)) + '円' + '×' + '{:,}'.format(int(amount)) + 'kg'
                l[UNIT_CODE_LIST_INDEX] = '2'

            elif unit_price == 2:
                l[UNIT_PRICE_LIST_INDEX] = total
                l[AMOUNT_LIST_INDEX] = 1
                l[REMARK_LIST_INDEX] = str(int(unit_price)) + '円' + '×' + '{:,}'.format(int(amount)) + 'kg'
                l[UNIT_CODE_LIST_INDEX] = '2'

            else:
                l[UNIT_PRICE_LIST_INDEX] = int(unit_price)
                l[AMOUNT_LIST_INDEX] = int(amount)

        elif pt_code == '10101':
            l[UNIT_PRICE_LIST_INDEX] = int(unit_price) + 15.0
            l[AMOUNT_LIST_INDEX] = int(amount)
            l[TOTAL_LIST_INDEX] = int(amount) * int(unit_price)
            l[UNIT_CODE_LIST_INDEX] = '0'

        elif pt_code == '03365':
            if unit_price == 7.7:
                kg_remake()

        elif pt_code == '03761':
            if unit_price == 1500:
                l[UNIT_PRICE_LIST_INDEX] = total
                l[AMOUNT_LIST_INDEX] = 1
                l[REMARK_LIST_INDEX] = '{:,}'.format(int(unit_price)) + '円' + '×' + str(amount) + '立米'
                l[UNIT_CODE_LIST_INDEX] = '2'

            else:
                l[UNIT_PRICE_LIST_INDEX] = int(unit_price)
                l[AMOUNT_LIST_INDEX] = int(amount)
                l[UNIT_CODE_LIST_INDEX] = 2

        #  単位がt(半角）だった場合
        elif unit == 't':
            kg_conversion()

        #  単位がt(全角）だった場合
        elif unit == 'ｔ':
            kg_conversion()

        #  単価が空で尚且つ、数量も空だった場合
        elif unit_price == '':
            if not amount:
                l[UNIT_PRICE_LIST_INDEX] = total
                l[AMOUNT_LIST_INDEX] = 1
                l[UNIT_CODE_LIST_INDEX] = 2

            elif int(amount) == 1:
                l[UNIT_PRICE_LIST_INDEX] = total
                l[AMOUNT_LIST_INDEX] = 1
                l[REMARK_LIST_INDEX] = str(amount) + str(unit)
                # l[UNIT_CODE_LIST_INDEX] = '単位を確認してください'
                l[UNIT_CODE_LIST_INDEX] = '2'

            else:
                l[UNIT_PRICE_LIST_INDEX] = total
                l[AMOUNT_LIST_INDEX] = 1
                l[REMARK_LIST_INDEX] = str(amount) + str(unit)



        ws.cell(column=1, row=i).value = l[STORE_CODE_LIST_INDEX]
        ws.cell(column=55, row=i).value = l[DAY_LIST_INDEX]
        ws.cell(column=68, row=i).value = l[PT_CODE_LIST_INDEX]
        ws.cell(column=77, row=i).value = l[UNIT_PRICE_LIST_INDEX]
        ws.cell(column=78, row=i).value = l[AMOUNT_LIST_INDEX]
        ws.cell(column=74, row=i).value = l[REMARK_LIST_INDEX]
        ws.cell(column=75, row=i).value = l[PRODUCT_CODE_LIST_INDEX]
        ws.cell(column=80, row=i).value = l[PRODUCT_NAM_LIST_INDEX]
        ws.cell(column=82, row=i).value = l[ITEM_CODE_LIST_INDEX]
        ws.cell(column=83, row=i).value = l[UNIT_CODE_LIST_INDEX]
        ws.cell(column=54, row=i).value = 2  # 契約タイプ区分

        emission_column = ['B', 'M', 'X', 'AI']
        zero_column = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'Y', 'Z',
                       'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AQ', 'AU',
                       'AV', 'AW', 'AZ', 'CA']
        one_column = ['BD', 'BE', 'BF', 'BG', 'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BX']

        for c in emission_column:
            column = openpyxl.utils.column_index_from_string(c)  # 列を数値に変換
            ws.cell(row=i, column=column).value = 0.00

        for c in zero_column:
            column = openpyxl.utils.column_index_from_string(c)
            ws.cell(row=i, column=column).value = 0

        for c in one_column:
            column = openpyxl.utils.column_index_from_string(c)
            ws.cell(row=i, column=column).value = 1

        i += 1
        row_count += 1

    max_row = ws.max_row + 1
    # import_data_formatの6行目以降の不要行を削除
    ws.delete_rows(idx=row_count + 6, amount=max_row - (row_count + 6))  # idx= 何行目から　amount= 何行分削除するか

    output_data = openpyxl.writer.excel.save_virtual_workbook(wb)

    return output_data



class ImportDataDetailView(LoginRequiredMixin, DetailView):
    model = ImportData
    template_name = 'app/import_data_detail.html'

    def get_context_data(self, **kwargs):
        import_data_pk = self.kwargs['pk']
        import_data = ImportData.objects.get(pk=import_data_pk)
        matched_data_pk = import_data.visually_matched.matched.pk
        matched_data = MatchedData.objects.get(pk=matched_data_pk)
        billing = BillingData.objects.filter(matched_id=matched_data_pk)

        context = super().get_context_data(**kwargs)
        context['import_data'] = import_data
        context['matched_data'] = matched_data
        if billing.exists():
            context['billing'] = billing
        context['status'] = IMPORT_DATA_OUTPUT_COMPLETED
        return context





